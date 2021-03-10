import openpyxl
path = "loction of the file"
workbook = openpyxl.load_workbook(path)                         # file location

sheet = workbook.active

rows = sheet.max_row
column = sheet.max_column

print(rows, column)
                                                                # read
for r in range(1,rows+1):
    for c in range(1, column+1):
        print(sheet.cell(rows=r,column=c).value)


                                                                # write

for r in range(1,rows+1):
    for c in range(1, column+1):
        sheet.cell(rows=r,column=c).value = "xyz"


workbook.save(path)